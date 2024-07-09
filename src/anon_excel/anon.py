import argparse
import logging
import os
from pathlib import Path
import pandas as pd
import re
import sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from anon_excel.calc_stats import (
    determine_distinct_students, paired_ttest)
import anon_excel.constants as const
from anon_excel.survey_files import find_survey_files, load_and_prepare_survey_data
from anon_excel.ranking_data import load_ranking_from_folder

log = logging.getLogger(__name__)


def get_parser() -> argparse.ArgumentParser:
    '''Setup a command line parser'''
    parser = argparse.ArgumentParser(
        description='''This app scans multiple sets of surveys. It offers an option
         to clean and store the survey data, and also an option to perform and store
         a T-test analysis. The T-test is only possible when both pre- and post-
         survey is available.
         Optionally personal information is removed.
        '''
    )
    parser.add_argument(
        'folder',
        help='Specify the folder with the excel report(s)')
    parser.add_argument(
        '-a', '--anonymize',
        action='count',
        required=False,
        default=0,
        help='Anonymize personal data (default = No)')
    parser.add_argument(
        '-c', '--color',
        action='store_true',
        required=False,
        default=False,
        help='Add colors in excel file with clean ranked data (default = No)')
    parser.add_argument(
        '-o', '--overwrite',
        action='store_true',
        required=False,
        default=False,
        help='Overwrite existing excel outputs (default = No)')
    parser.add_argument(
        '-s', '--strip',
        action='store_true',
        required=False,
        default=False,
        help='Strip leading s-char from s-number (default = No)')
    parser.add_argument(
        '-t', '--ttest',
        action='store_true',
        required=False,
        default=False,
        help='Perform T-test calculation (default = No)')
    parser.add_argument(
        '-x', '--clean',
        action='store_true',
        required=False,
        default=False,
        help='Save cleaned data (default = No)')

    return parser


def remove_previous_results(files: list[Path], do_overwrite: bool, which_output: str) -> bool:
    if not files:
        return True

    prev = [Path(p).name for p in files]
    if do_overwrite:
        log.info(f'Trying to remove previous {which_output} results: \n{prev}')
        for f in files:
            os.remove(f)
        return True

    log.error(f'Output {which_output} data already exists.\n'
              'Use --overwrite to force removal and recalculation')

    return False


def check_remove_all_outputs(folder: Path, clean: bool, ttest: bool, overwrite: bool) -> bool:
    for check, rem in zip([const.ANALYSIS_OUTPUT_BASE, const.CLEANED_OUTPUT_BASE], [ttest, clean or ttest]):
        if not rem:
            continue
        cur_fol = folder / check
        prev = list(cur_fol.glob(f'{check}*.xlsx'))
        if not remove_previous_results(prev, which_output=check, do_overwrite=overwrite):
            return False

    return True


def check_create_out_folder(folder: Path):
    '''Make sure folder exists'''
    if folder.exists() and folder.is_dir():
        return True

    folder.mkdir()


def write_to_excel(filename: Path, sheets: tuple[list[pd.DataFrame], str]) -> None:
    with pd.ExcelWriter(filename, engine='xlsxwriter') as writer:
        for df, sheetname in sheets:
            df.to_excel(writer, sheet_name=sheetname, index=False)


def validate_input(args) -> tuple[Path, str]:
    folder = Path(args.folder)
    if not folder.exists:
        log.error('Folder {args.folder} does not exist')
        sys.exit(1)

    id_column = const.DEFAULT_STUDENT_COLUMN

    if not check_remove_all_outputs(folder,
                                    clean=args.clean,
                                    ttest=args.ttest,
                                    overwrite=args.overwrite):
        sys.exit()
    return folder, id_column


def determine_survey_data_name(survey_file: Path, sequence_nr: int) -> str:
    '''Extract a sequence ID from filename. The pattern looked for is a single
       `(n-m)` at the end of the filename where n, m are integer values, fe. `(1-89)`.
       If this cannot be found the `sequence_nr` will be used instead to
       generate a filename.
       The name generated will then be either:
       1. data_survey_(n-m)
       2. data_survey_<sequence_nr>
    '''
    patt = f'{sequence_nr:02}'

    regex = r'.*(\(\d+-\d+\))'
    m = re.match(regex, survey_file.name)
    if m and (len(m.groups()) == 1):
        patt = m.group(1)

    return f'{const.DATA_OUTPUT_BASENAME}_{patt}'


def clean_and_save_survey_data(folder: Path,
                               pre_file: Path, post_file: Path,
                               df_pre: pd.DataFrame, df_post: pd.DataFrame,
                               seq_nr: int,
                               anonymize: int) -> Path:
    ''' Save survey data to excel. Only keep relevant columns.
        Depending on the command line parameter `anonymize` the following happens:
        `anonymize=0` : student ID column is retained, no anonymized data in output
        `anonymize=1` : student ID column is retained, also anonymized data in output
        `anonymize=2` : student ID column is removed, only anonymized data in output
    '''
    out_folder = folder / const.CLEANED_OUTPUT_BASE
    check_create_out_folder(out_folder)
    clean_output = out_folder / \
        f'''{const.CLEANED_OUTPUT_BASE}_{determine_survey_data_name(
            pre_file, sequence_nr=seq_nr)}.xlsx'''

    # filter out the sensitive columns and prepare for output
    cols_to_drop = const.DROP_COLUMNS
    if anonymize == 0:  # 0 == drop anonymized data
        log.info(f'Do not save anonymized data, {anonymize=}')
        cols_to_drop = [*const.DROP_COLUMNS, const.ANONYMOUS_ID]
    if anonymize == 2:  # 2 == drop sensitive data
        log.info(f'Removing sensitive data, {anonymize=}')
        cols_to_drop = [*const.DROP_COLUMNS, const.DEFAULT_STUDENT_COLUMN]
    remain_columns = [col for col in df_pre.columns if col not in cols_to_drop]
    clean_data = [(df_pre[remain_columns], const.CLEAN_SHEET_PRE_SURVEY)]
    if post_file.name:
        remain_columns = [
            col for col in df_post.columns if col not in cols_to_drop]
        clean_data.append((df_post[remain_columns], const.CLEAN_SHEET_POST_SURVEY))

    log.info(f'Writing cleaned data to "{clean_output}"')
    write_to_excel(clean_output, sheets=clean_data)

    return clean_output


def ttest_and_save(folder: Path, pre_file: Path,
                   df_pre: pd.DataFrame, df_post: pd.DataFrame,
                   seq_nr: int):
    log.info('Calculating paired T-test from Pre- and Post survey')
    df_pairs, _, df_legend, df_bf, df_af, df_stud_pairs = paired_ttest(
        df_pre, df_post, id_column=const.ANONYMOUS_ID)

    out_folder = folder / const.ANALYSIS_OUTPUT_BASE
    check_create_out_folder(out_folder)
    excel_output = out_folder / \
        f'{const.ANALYSIS_OUTPUT_BASE}_{
            determine_survey_data_name(pre_file, seq_nr)}.xlsx'
    ttest_output = [(df_pairs, 'Paired T-test'),
                    (df_stud_pairs, 'Students T-test'),
                    (df_bf, 'Pre-questions'),
                    (df_af, 'Post-questions'),
                    (df_legend, 'Question legend')]

    log.info(f'Writing analysis result to "{excel_output}"')
    write_to_excel(excel_output, sheets=ttest_output)

# Function to find the column index of 'id_column'


def find_id_column_index(worksheet, column_name: str):
    for cell in worksheet[1]:  # Iterate over the first row (header)
        if cell.value == column_name:
            return cell.col_idx - 1  # Return zero-based index


def colorize_excel(excel_file: Path,
                   df_pre: pd.DataFrame, df_post: pd.DataFrame,
                   id_column: str):
    book = load_workbook(excel_file)

    stud_common, studs_before_only, stud_after_only = determine_distinct_students(
        df_pre, df_post, id_column)

    # Define fills
    fill_green = PatternFill(start_color="CCFFCC",
                             end_color="CCFFCC", fill_type="solid")
    fill_blue = PatternFill(start_color="CCCCFF", end_color="CCCCFF", fill_type="solid")
    fill_red = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    # Get the worksheets
    ws_pre = book[const.CLEAN_SHEET_PRE_SURVEY]
    ws_post = book[const.CLEAN_SHEET_POST_SURVEY]

    # get index of the id_column
    id_col_idx_pre = find_id_column_index(ws_pre, id_column)
    id_col_idx_post = find_id_column_index(ws_post, id_column)

    # Apply conditional formatting for survey_pre
    for row in ws_pre.iter_rows(min_row=2, max_row=ws_pre.max_row, min_col=1, max_col=ws_pre.max_column):
        cell_value = row[id_col_idx_pre].value
        if cell_value in stud_common:
            for cell in row:
                cell.fill = fill_green
        elif cell_value in studs_before_only:
            for cell in row:
                cell.fill = fill_blue

    # Apply conditional formatting for survey_post
    for row in ws_post.iter_rows(min_row=2, max_row=ws_post.max_row, min_col=1, max_col=ws_post.max_column):
        cell_value = row[id_col_idx_post].value
        if cell_value in stud_common:
            for cell in row:
                cell.fill = fill_green
        elif cell_value in stud_after_only:
            for cell in row:
                cell.fill = fill_red

    # Save the updated Excel file
    book.save(excel_file)


def main():
    args = get_parser().parse_args()

    folder, id_column = validate_input(args)

    surveys = find_survey_files(folder, allow_missing_post=args.clean)
    if not surveys:
        log.error('No survey files found, quitting')
        sys.exit()

    if not args.clean and not args.ttest:
        log.info('No cleaning or t-test requested, nothing to do, quitting')
        sys.exit()

    # init ranking lookup
    load_ranking_from_folder(args.folder)

    # start processing
    for seq_nr, (pre_file, post_file) in enumerate(surveys, start=1):
        log.info('Initiating analysis')
        log.info(f'Pre-survey file: "{pre_file}"')
        df_pre = load_and_prepare_survey_data(pre_file, id_column, args.strip)
        if post_file.name:
            log.info(f'Post-survey file: "{post_file}"')
            df_post = load_and_prepare_survey_data(post_file, id_column, args.strip)
        else:
            log.info('No accompanying post file')
            if args.ttest:
                log.info('Skipping T-test')

        if args.clean:
            clean_output = clean_and_save_survey_data(
                folder, pre_file, post_file, df_pre, df_post,
                seq_nr,
                args.anonymize)

            # optionally apply styles/colors
            if args.color:
                sel_col = const.ANONYMOUS_ID
                if args.anonymize == 0:
                    sel_col = id_column
                colorize_excel(clean_output, df_pre, df_post, sel_col)

        # T-test is only possible whith both pre- and post_survey files
        if args.ttest and post_file.name:
            ttest_and_save(folder, pre_file, df_pre, df_post, seq_nr)


if __name__ == '__main__':
    main()
